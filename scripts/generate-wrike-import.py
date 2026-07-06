"""
Generate a Wrike-ready Excel workbook from the QA findings dataset.

Input : scripts/qa-findings.json  (produced by `node scripts/export-findings.ts`)
Output: wrike-import.xlsx          (import via Wrike > Import from Excel)

Each finding becomes one row / one Wrike task, structured to match the
existing O2 Wrike issue layout:
  Title            -> "ISSUE | <title>"
  Importance       -> High / Normal / Low  (Wrike built-in)
  Status           -> New
  Assigned To      -> likely developer (author of the offending code)
  Description      -> Steps to reproduce / Expected Result / Current results
  + custom columns : Severity, Product, Repository, Branch, Location, Finding ID
"""
import json
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DATA = os.path.join(HERE, "qa-findings.json")
WEB_DATA = os.path.join(HERE, "website-findings.json")
OUT = os.path.join(ROOT, "wrike-import.xlsx")

# Map severities to Wrike's 3-level Importance field.
IMPORTANCE = {
    "critical": "High",
    "high": "High",
    "medium": "Normal",
    "low": "Low",
    "info": "Low",
}

# Normalise audit "author" strings to the names used in the Wrike workspace.
NAME_FIXES = {
    "Peter Njuguna": "Peter Ndungu",  # git: xcixor <pndungu54@gmail.com>
}


def clean_assignee(author: str) -> str:
    author = re.sub(r"\([^)]*\)", "", author)          # drop "(xcixor)" etc.
    parts = re.split(r"[/,]", author)
    names = []
    for p in parts:
        p = p.strip()
        if not p or "branch owner" in p.lower():
            continue
        names.append(NAME_FIXES.get(p, p))
    return ", ".join(dict.fromkeys(names))              # dedupe, keep order


def build_description(f: dict) -> str:
    """Developer-friendly layout: lead with WHAT and WHERE, then repro,
    then expected vs current. Plain labelled sections so a developer can
    scan the issue and jump straight to the file:line."""
    return (
        f">> IDENTIFIED ISSUE\n{f['issue']}\n"
        f"\n"
        f">> WHERE TO LOOK (file : line)\n{f['location']}\n"
        f"\n"
        f">> EXPECTED RESULT\n{f['expected']}\n"
        f"\n"
        f">> CURRENT RESULT\n{f['actual']}\n"
        f"\n"
        f"----------------------------------------\n"
        f"Severity: {f['severity'].upper()}  |  Product: {f['product']}  |  "
        f"Repo: {f['repo']}  |  Branch: {f['branch']}  |  Ref: {f['id']}"
    )


# Location placed as the 2nd column so developers can see WHERE the issue is
# at a glance in the spreadsheet, before drilling into the description.
COLUMNS = [
    ("Title", 58),
    ("Location (file : line)", 46),
    ("Severity", 11),
    ("Importance", 11),
    ("Status", 10),
    ("Assigned To", 22),
    ("Product", 9),
    ("Repository", 34),
    ("Branch", 13),
    ("Finding ID", 13),
    ("Identified Issue / Details", 100),
]

HEADER_FILL = PatternFill("solid", fgColor="F97316")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def write_sheet(ws, findings):
    ws.append([c[0] for c in COLUMNS])
    for i, (name, width) in enumerate(COLUMNS, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = width
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
    for f in sorted(findings, key=lambda x: (order.get(x["severity"], 9), x["id"])):
        ws.append([
            f"ISSUE | {f['title']}",
            f["location"],
            f["severity"].capitalize(),
            IMPORTANCE.get(f["severity"], "Normal"),
            "New",
            clean_assignee(f["author"]),
            f["product"],
            f["repo"],
            f["branch"],
            f["id"],
            build_description(f),
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP_TOP


# ---------------------------------------------------------------------------
# Website findings (live-site E2E QA) — different shape from repo findings:
# site / page / category / screenshot instead of product / repo / branch.
# ---------------------------------------------------------------------------
WEB_COLUMNS = [
    ("Title", 58),
    ("Location", 46),
    ("Severity", 11),
    ("Importance", 11),
    ("Status", 10),
    ("Site", 8),
    ("Category", 15),
    ("Page", 34),
    ("Finding ID", 15),
    ("Screenshot", 34),
    ("Identified Issue / Details", 100),
]


def build_web_description(f: dict) -> str:
    return (
        f">> IDENTIFIED ISSUE / HOW TO REPRODUCE\n{f['issue']}\n"
        f"\n"
        f">> WHERE (page / element)\n{f['page']}  |  {f['location']}\n"
        f"\n"
        f">> EXPECTED RESULT\n{f['expected']}\n"
        f"\n"
        f">> CURRENT RESULT\n{f['actual']}\n"
        f"\n"
        f">> SCREENSHOT EVIDENCE\n{f['screenshot']}\n"
        f"\n"
        f"----------------------------------------\n"
        f"Severity: {f['severity'].upper()}  |  Site: {f['site']}  |  "
        f"Category: {f['category']}  |  Ref: {f['id']}"
    )


def write_web_sheet(ws, findings):
    ws.append([c[0] for c in WEB_COLUMNS])
    for i, (name, width) in enumerate(WEB_COLUMNS, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = width
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
    for f in sorted(findings, key=lambda x: (order.get(x["severity"], 9), x["id"])):
        ws.append([
            f"ISSUE | {f['title']}",
            f["location"],
            f["severity"].capitalize(),
            IMPORTANCE.get(f["severity"], "Normal"),
            "New",
            f["site"],
            f["category"],
            f["page"],
            f["id"],
            f["screenshot"],
            build_web_description(f),
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP_TOP


def main():
    with open(DATA, encoding="utf-8") as fh:
        payload = json.load(fh)
    findings = payload["qaFindings"]

    wb = Workbook()
    write_sheet(wb.active, findings)
    wb.active.title = "All Findings"

    products = sorted({f["product"] for f in findings})
    for product in products:
        ws = wb.create_sheet(product)
        write_sheet(ws, [f for f in findings if f["product"] == product])

    # Website findings sheet(s)
    web_count = 0
    if os.path.exists(WEB_DATA):
        with open(WEB_DATA, encoding="utf-8") as fh:
            web_payload = json.load(fh)
        web_findings = web_payload["websiteFindings"]
        web_count = len(web_findings)
        write_web_sheet(wb.create_sheet("Website"), web_findings)

    out_path = OUT
    try:
        wb.save(OUT)
    except PermissionError:
        import time
        out_path = os.path.join(ROOT, f"wrike-import-{time.strftime('%Y%m%d-%H%M%S')}.xlsx")
        wb.save(out_path)
        print(f"[warn] {OUT} is locked (open in Excel?). Saved to {out_path} instead.")

    counts = ", ".join(
        f"{sum(1 for f in findings if f['product'] == p)} {p}" for p in products
    )
    print(f"Wrote {len(findings)} repo findings ({counts}) + {web_count} website findings -> {out_path}")


if __name__ == "__main__":
    main()
